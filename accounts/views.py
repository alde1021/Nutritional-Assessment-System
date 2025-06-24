from django.shortcuts import get_object_or_404, render, HttpResponse, redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,login,logout
from django.contrib.auth.decorators import login_required
from accounts.forms import ProfileForm
from accounts.models import Student
from django.contrib import messages
from django.db.models import Count
from collections import Counter
from collections import defaultdict
from adminpanel.models import AssessmentSchedule
from adminpanel.models import CustomUser
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import openpyxl
from django.http import HttpResponse
from io import BytesIO
from django.db.models import Q
# Create your views here.

# Fixed order of categories
NUTRITIONAL_STATUS_CATEGORIES = ['Severely Wasted', 'Wasted', 'Normal', 'Overweight', 'Obese']
HEIGHT_FOR_AGE_CATEGORIES = ['Severely Stunted', 'Stunted', 'Normal', 'Tall']

def student_list(request):
     mem = Student.objects.all()
     return render(request, 'accounts/index.html', {'mem':mem})

def index(request):
    user = request.user  # ✅ Define 'user' before using it
    mem = Student.objects.all()  # Get all student records
    students = Student.objects.filter(user = user)
    if request.method == 'POST':
        form = ProfileForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('index')  # Or 'index' if that's the correct URL name
        else:
            print(form.errors)
    else:
        form = ProfileForm()

    context = {
        'form': form,
        'mem': mem,
    }

    return render(request, 'accounts/index.html', context)

def HomePage(request):
    return render (request,'accounts/home.html')

def LoginPage(request):
    if request.method=='POST':
        username=request.POST.get('username')
        password=request.POST.get('password')

        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
            else:
                messages.error(request, "Invalid username or password.")
        else:
            messages.error(request, "Please enter both username and password.")

    return render (request,'accounts/login.html')

def LogoutPage(request):
    logout(request)
    return redirect('login')


def dashboard(request):

    # 1. Distinct values for dropdowns
    assessment_types = Student.objects.values_list('assessment_type', flat=True).distinct()
    school_years = Student.objects.values_list('school_year', flat=True).distinct()
    grade_levels = Student.objects.values_list('grade_level', flat=True).distinct()
    sections = Student.objects.values_list('section', flat=True).distinct()

    # 2. Get selected values from request
    selected_type = request.GET.get('assessment_type')
    selected_year = request.GET.get('school_year')
    selected_grade = request.GET.get('grade_level')
    selected_section = request.GET.get('section')

    # 3. Filter students based on selected inputs
    students = Student.objects.all()
    if selected_type:
        students = students.filter(assessment_type=selected_type)
    if selected_year:
        students = students.filter(school_year=selected_year)
    if selected_grade:
        students = students.filter(grade_level=selected_grade)
    if selected_section:
        students = students.filter(section=selected_section)

    # 4. Counts for summary
    total_students = students.count()
    total_male = students.filter(gender='Male').count()
    total_female = students.filter(gender='Female').count()

    # Categorize dynamically
    bmi_counts_male = Counter(s.nutritional_status() for s in students if s.gender == 'Male')
    bmi_counts_female = Counter(s.nutritional_status() for s in students if s.gender == 'Female')

    height_counts_male = Counter(s.height_for_age() for s in students if s.gender == 'Male')
    height_counts_female = Counter(s.height_for_age() for s in students if s.gender == 'Female')
   # Build chart data
    labels = sorted(set(bmi_counts_male.keys()) | set(bmi_counts_female.keys()))
    data = [bmi_counts_male.get(lbl, 0) + bmi_counts_female.get(lbl, 0) for lbl in labels]

    return render(request, 'accounts/dashboard.html', {
        'students': students,
        'assessment_types': assessment_types,
        'school_years': school_years,
        'grade_levels': grade_levels,
        'sections': sections,
        'selected': {
            'assessment_type': selected_type,
            'school_year': selected_year,
            'grade_level': selected_grade,
            'section': selected_section,
        },
        'bmi_counts_male': bmi_counts_male,
        'bmi_counts_female': bmi_counts_female,
        'height_counts_male': height_counts_male,
        'height_counts_female': height_counts_female,
        'total_students': total_students,
        'total_male': total_male,
        'total_female': total_female,
        'labels': labels,
        'data': data,
        })

def Add_student(request):
    user = request.user
    school_years = AssessmentSchedule.objects.values_list('school_year', flat=True).distinct()  
    mem = Student.objects.filter(user=user)  # Show only the current user's students
    if request.method == 'POST':
        form = ProfileForm(request.POST)
        if form.is_valid():
            student = form.save(commit=False)
            student.user = user
            student.grade_level = user.grade_level
            student.section = user.section
            student.save()
            return redirect('add')  # Or your target URL
        else:
            print(form.errors)
    else:
        form = ProfileForm()

    context = {
        'form': form,
        'mem': mem,
        'school_years': school_years,
    }

    return render(request, 'accounts/add.html', context)
# updpate view
def update_student(request, id_no):
    student = get_object_or_404(Student, id_no=id_no)
    if request.method == 'POST':
        student.first = request.POST.get('first', student.first)
        student.last = request.POST.get('last', student.last)
        student.middle = request.POST.get('middle', student.middle)
        student.birthday = request.POST.get('birthday', student.birthday)
        student.weight = request.POST.get('weight', student.weight)
        student.height = request.POST.get('height', student.height)
        student.gender = request.POST.get('gender', student.gender)
        student.save()
        messages.success(request, f"{student.first} {student.last} was successfully updated.")
    return redirect('index')
# delete view

def delete_student(request, id_no):
    mem = get_object_or_404(Student, id_no=id_no)

    if request.method == 'POST':
        mem.delete()
        messages.success(request, f"Student {mem.first} {mem.last} has been deleted.")
        return redirect('index')  # Adjust to your actual student list view name

    return render(request, 'accounts/delete.html', {'mem': mem})

def report_view(request):
    students = []
    nutrition_counts = {}
    height_counts = {}

    grade_level = request.GET.get('grade_level')
    section = request.GET.get('section')
    assessment_type = request.GET.get('assessment_type')
    school_year = request.GET.get('school_year')
    prepared_by = request.user.get_full_name() or request.user.username
    # 🟩 Get filtered students
    if grade_level and section and assessment_type and school_year:
        students = Student.objects.filter(
            grade_level=grade_level,
            section=section,
            assessment_type=assessment_type,
            school_year=school_year
        )
     # 🟩 Generate summaries
    bmi_categories = [s.nutritional_status() for s in students]
    height_categories = [s.height_for_age() for s in students]

    nutrition_counts = dict(Counter(bmi_categories))
    height_counts = dict(Counter(height_categories))
        # 🟩 Get assessment schedule record
    schedule = AssessmentSchedule.objects.filter(school_year=school_year).first()

        # 🟩 If Export is requested
    if 'export' in request.GET:
        if students.exists():
            student_sample = students.first()
            return export_to_excel(
                    students=students,
                    school_name=schedule.school_name if schedule else "N/A",
                    city=schedule.city if schedule else "N/A",
                    school_year=student_sample.school_year,
                    grade_level=student_sample.grade_level,
                    section=student_sample.section,
                    date_of_weighing=student_sample.date_of_weighing,
                    prepared_by=prepared_by,
                    assessment_type=assessment_type   
                )
        else:
                return HttpResponse("No students found to export.")
    # Normal render
    

    return render(request, 'accounts/reports.html', {
        'students': students,
        'grade_levels': Student.objects.values_list('grade_level', flat=True).distinct(),
        'sections': Student.objects.values_list('section', flat=True).distinct(),
        'school_years': Student.objects.values_list('school_year', flat=True).distinct(),
        'selected_grade': grade_level,
        'selected_section': section,
        'selected_assessment': assessment_type,
        'selected_sy': school_year,
        'nutrition_counts': nutrition_counts,
        'height_counts': height_counts,
    })

def export_to_excel(students, school_name, city, school_year, grade_level, section, date_of_weighing, prepared_by, assessment_type):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Nutritional Report"

    # Styles
    bold_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    gray_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

    # Header
    sheet.merge_cells('A1:K1')
    sheet['A1'] = "NUTRITIONAL STATUS REPORT"
    sheet['A1'].font = Font(bold=True, size=14)
    sheet['A1'].alignment = center

    sheet.merge_cells('A2:K2')
    sheet['A2'] = school_name
    sheet['A2'].alignment = center

    sheet.merge_cells('A3:K3')
    sheet['A3'] = city
    sheet['A3'].alignment = center

    sheet.merge_cells('A4:K4')
    sheet['A4'] = f"SY {school_year} - {assessment_type}"
    sheet['A4'].alignment = center

    sheet['A5'] = f"Date of Weighing: {date_of_weighing.strftime('%B %d, %Y')}"
    sheet['A5'].font = bold_font
    sheet['A5'].alignment = Alignment(horizontal='left')

    sheet['K5'] = f"Grade {grade_level}: {section}"
    sheet['K5'].font = bold_font
    sheet['K5'].alignment = Alignment(horizontal='right')

    # Table headers
    headers = ["No.", "Names", "Birthday", "Weight (kg)", "Height (m)", "Sex", "Height² (m²)", "Age (y,m)", "BMI", "Nutritional Status", "Height-For-Age"]
    sheet.append(headers)
    header_row = sheet.max_row
    for col_num, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_num)
        cell.font = bold_font
        cell.alignment = center
        cell.border = border

    # Student data
    for idx, student in enumerate(students, start=1):
        row = [
            idx,
            f"{student.last.upper()}, {student.first.upper()}",
            student.birthday.strftime("%m/%d/%Y"),
            student.weight,
            student.height,
            student.gender,
            round(student.height ** 2, 4),
            student.age_years_months(),
            round(student.bmi(), 2),
            student.nutritional_status(),
            student.height_for_age()
        ]
        sheet.append(row)
        for col_num in range(1, len(row) + 1):
            cell = sheet.cell(row=sheet.max_row, column=col_num)
            cell.alignment = center
            cell.border = border

    # Auto-fit columns
    for col in sheet.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # BMI and HFA Summary Side-by-Side
    bmi_start = sheet.max_row + 2
    hfa_start_col = 7

    # BMI Header
    sheet.merge_cells(start_row=bmi_start, start_column=1, end_row=bmi_start, end_column=4)
    cell = sheet.cell(row=bmi_start, column=1)
    cell.value = "Body Mass Index"
    cell.font = bold_font
    cell.alignment = center
    cell.fill = yellow_fill

    sheet.cell(row=bmi_start+1, column=1).value = "No. of Cases"
    sheet.cell(row=bmi_start+1, column=1).font = bold_font
    sheet.cell(row=bmi_start+1, column=1).fill = yellow_fill
    for i, label in enumerate(["M", "F", "T"], start=2):
        cell = sheet.cell(row=bmi_start+1, column=i)
        cell.value = label
        cell.font = bold_font
        cell.alignment = center
        cell.border = border

    # HFA Header
    sheet.merge_cells(start_row=bmi_start, start_column=hfa_start_col, end_row=bmi_start, end_column=hfa_start_col+3)
    cell = sheet.cell(row=bmi_start, column=hfa_start_col)
    cell.value = "Height-for-Age (HFA)"
    cell.font = bold_font
    cell.alignment = center
    cell.fill = yellow_fill

    sheet.cell(row=bmi_start+1, column=hfa_start_col).value = "No. of Cases"
    sheet.cell(row=bmi_start+1, column=hfa_start_col).font = bold_font
    sheet.cell(row=bmi_start+1, column=hfa_start_col).fill = yellow_fill
    for i, label in enumerate(["M", "F", "T"], start=1):
        cell = sheet.cell(row=bmi_start+1, column=hfa_start_col+i)
        cell.value = label
        cell.font = bold_font
        cell.alignment = center
        cell.border = border

    bmi_categories = ["Severely Wasted", "Wasted", "Normal", "Overweight", "Obese"]
    hfa_categories = ["Sev. Stunted", "Stunted", "Normal", "Tall"]

    for i in range(max(len(bmi_categories), len(hfa_categories))):
        if i < len(bmi_categories):
            cat = bmi_categories[i]
            m = sum(1 for s in students if s.gender == 'Male' and s.nutritional_status() == cat)
            f = sum(1 for s in students if s.gender == 'Female' and s.nutritional_status() == cat)
            row = [cat, m, f, m + f]
            for col, val in enumerate(row, start=1):
                cell = sheet.cell(row=bmi_start + 2 + i, column=col)
                cell.value = val
                cell.alignment = center
                cell.border = border
                if col == 1:
                    cell.font = bold_font
                    cell.fill = gray_fill

        if i < len(hfa_categories):
            cat = hfa_categories[i]
            m = sum(1 for s in students if s.gender == 'Male' and s.height_for_age() == cat)
            f = sum(1 for s in students if s.gender == 'Female' and s.height_for_age() == cat)
            row = [cat, m, f, m + f]
            for col, val in enumerate(row, start=hfa_start_col):
                cell = sheet.cell(row=bmi_start + 2 + i, column=col)
                cell.value = val
                cell.alignment = center
                cell.border = border
                if col == hfa_start_col:
                    cell.font = bold_font
                    cell.fill = gray_fill

    # Prepared by
    final_row = sheet.max_row + 2
    sheet.merge_cells(start_row=final_row, start_column=1, end_row=final_row, end_column=5)
    prepared = sheet.cell(row=final_row, column=1)
    prepared.value = f"Prepared by: {prepared_by}"
    prepared.font = bold_font
    prepared.alignment = center

    # Export
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Nutritional_Report_{grade_level}_{section}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)
    return response

def student_progress_search(request):
    query = request.GET.get('last_name')  # still using 'last_name' as the GET parameter
    records = Student.objects.none()
    student = None

    if query:
        if ',' in query:
            # Match full name format: "Last, First"
            parts = [p.strip() for p in query.split(',', 1)]
            if len(parts) == 2:
                last, first = parts
                records = Student.objects.filter(last__iexact=last, first__iexact=first).order_by('grade_level')
        else:
            # Partial match on last name only
            records = Student.objects.filter(last__icontains=query).order_by('grade_level')

        student = records.first()

    return render(request, 'accounts/progress_tracker.html', {
        'records': records,
        'student': student,
        'query': query,
    })